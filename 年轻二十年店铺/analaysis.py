"""
门店日记账分析 —— 每个月每个项目的每日平均成本
"""

import os
import sys
import datetime

import openpyxl
import pandas as pd

sys.stdout.reconfigure(encoding="utf-8")

# ── 列映射（按 Excel 列序号，1 起）───────────────────────────────────────────

COL_DATE = 1

COL_MAP = {
    "堂食收入":     2,
    "堂食份数":     3,
    "外卖收入":     5,
    "外卖份数":     6,
    "团购收入":     8,
    "团购份数":     9,
    "外卖退单":    11,
    "排队退单":    12,
    "活动折扣支出": 13,
    "合计营收":    14,
    # ── 食材成本 ──
    "上班(食材)":  15,
    "蔬菜/促销":   16,
    "调料/牛腩等": 17,
    "配料/黄皮酱": 18,
    "小料":        19,
    "其他消耗品":  20,
    # ── 运营成本 ──
    "推广支出":    21,
    "综合水电/包装": 22,
    "管理支出":    23,
    "增值费用":    24,
    "租赁房租":    25,
    "员工薪酬":    26,
    "水电费":      27,
    "其他费用":    28,
}

# 成本项（用于"每日平均成本"统计）
COST_COLS = [
    "上班(食材)", "蔬菜/促销", "调料/牛腩等", "配料/黄皮酱", "小料", "其他消耗品",
    "推广支出", "综合水电/包装", "管理支出", "增值费用",
    "租赁房租", "员工薪酬", "水电费", "其他费用",
]

MONTH_NAMES = {
    1: "1月", 2: "2月", 3: "3月", 4: "4月",
    5: "5月", 6: "6月", 7: "7月", 8: "8月",
    9: "9月", 10: "10月", 11: "11月", 12: "12月",
}


def read_sheet(ws) -> pd.DataFrame:
    """
    将一个月的工作表读取为 DataFrame，
    只保留 A 列是 datetime 且有实际营收的行。
    """
    rows = []
    for row in ws.iter_rows(
        min_row=7, max_row=ws.max_row, values_only=True
    ):
        date_val = row[COL_DATE - 1]
        if not isinstance(date_val, datetime.datetime):
            continue

        record = {"日期": date_val.date()}
        for name, col_idx in COL_MAP.items():
            val = row[col_idx - 1]
            record[name] = float(val) if isinstance(val, (int, float)) else 0.0

        rows.append(record)

    if not rows:
        return pd.DataFrame()

    return pd.DataFrame(rows)


def analyze_month(df: pd.DataFrame, month_name: str) -> pd.DataFrame:
    """
    计算该月：
    - 总计
    - 营业天数（有堂食或外卖收入的天数）
    - 每日平均成本（各支出项 / 营业天数）
    """
    if df.empty:
        return pd.DataFrame()

    operating_days = int(
        ((df["堂食收入"] > 0) | (df["外卖收入"] > 0)).sum()
    )
    if operating_days == 0:
        return pd.DataFrame()

    result_rows = []
    for col in COST_COLS:
        total = df[col].sum()
        daily_avg = total / operating_days if operating_days > 0 else 0
        result_rows.append({
            "月份":       month_name,
            "项目":       col,
            "月合计(元)": round(total, 2),
            "营业天数":   operating_days,
            "日均成本(元)": round(daily_avg, 2),
        })

    return pd.DataFrame(result_rows)


def main():
    xlsx_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "门店日记账.xlsx"
    )

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_names = wb.sheetnames  # ['1月', '2月', ...]

    all_results = []

    for idx, sheet_name in enumerate(sheet_names, 1):
        month_label = MONTH_NAMES.get(idx, sheet_name)
        ws = wb[sheet_name]
        df = read_sheet(ws)

        if df.empty:
            continue

        operating_days = int(
            ((df["堂食收入"] > 0) | (df["外卖收入"] > 0)).sum()
        )
        if operating_days == 0:
            continue

        result = analyze_month(df, month_label)
        all_results.append(result)

    if not all_results:
        print("未找到有效数据")
        return

    combined = pd.concat(all_results, ignore_index=True)
    save_csv(combined)

    # ── 输出：按月分组打印 ────────────────────────────────────────────────────
    sep = "─" * 72
    for month_label, group in combined.groupby("月份", sort=False):
        operating_days = group["营业天数"].iloc[0]
        total_cost = group["月合计(元)"].sum()
        daily_total = group["日均成本(元)"].sum()

        print(f"\n{'═' * 72}")
        print(f"  {month_label}    营业天数: {operating_days} 天    月支出合计: {total_cost:,.2f} 元    日均总支出: {daily_total:,.2f} 元")
        print(sep)
        print(f"  {'项目':<14}  {'月合计(元)':>12}  {'营业天数':>8}  {'日均成本(元)':>12}")
        print(sep)

        food_total = 0.0
        ops_total  = 0.0

        food_items = ["上班(食材)", "蔬菜/促销", "调料/牛腩等", "配料/黄皮酱", "小料", "其他消耗品"]
        ops_items  = ["推广支出", "综合水电/包装", "管理支出", "增值费用", "租赁房租", "员工薪酬", "水电费", "其他费用"]

        print(f"  ── 食材成本 ──")
        for _, row in group[group["项目"].isin(food_items)].iterrows():
            if row["月合计(元)"] == 0:
                continue
            print(f"  {row['项目']:<14}  {row['月合计(元)']:>12,.2f}  {row['营业天数']:>8}  {row['日均成本(元)']:>12,.2f}")
            food_total += row["月合计(元)"]

        food_daily = round(food_total / operating_days, 2)
        print(f"  {'  食材小计':<14}  {food_total:>12,.2f}  {operating_days:>8}  {food_daily:>12,.2f}")

        print(f"  ── 运营成本 ──")
        for _, row in group[group["项目"].isin(ops_items)].iterrows():
            if row["月合计(元)"] == 0:
                continue
            print(f"  {row['项目']:<14}  {row['月合计(元)']:>12,.2f}  {row['营业天数']:>8}  {row['日均成本(元)']:>12,.2f}")
            ops_total += row["月合计(元)"]

        ops_daily = round(ops_total / operating_days, 2)
        print(f"  {'  运营小计':<14}  {ops_total:>12,.2f}  {operating_days:>8}  {ops_daily:>12,.2f}")
        print(sep)
        print(f"  {'合计':<14}  {total_cost:>12,.2f}  {operating_days:>8}  {daily_total:>12,.2f}")

    # ── 汇总表（所有月份横向对比）────────────────────────────────────────────
    print(f"\n\n{'═' * 72}")
    print("  各月日均成本汇总对比")
    print(sep)

    pivot = combined.pivot_table(
        index="项目", columns="月份", values="日均成本(元)", aggfunc="sum"
    )
    pivot = pivot.reindex(COST_COLS)
    pivot = pivot.dropna(how="all")

    col_width = 10
    months_with_data = [c for c in MONTH_NAMES.values() if c in pivot.columns]
    header = f"  {'项目':<14}" + "".join(f"  {m:>{col_width}}" for m in months_with_data)
    print(header)
    print(sep)

    for item, row in pivot.iterrows():
        vals = "".join(
            f"  {row[m]:>{col_width},.1f}" if m in row and pd.notna(row[m]) else f"  {'—':>{col_width}}"
            for m in months_with_data
        )
        print(f"  {item:<14}{vals}")

    print(sep)
    daily_totals = pivot.sum()
    totals_str = "".join(
        f"  {daily_totals.get(m, 0):>{col_width},.1f}" for m in months_with_data
    )
    print(f"  {'日均合计':<14}{totals_str}")
    print()


def save_csv(combined: pd.DataFrame):
    """把结果保存为 UTF-8 BOM CSV，方便用 Excel 直接打开。"""
    out_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "日均成本分析.csv"
    )
    combined.to_csv(out_path, index=False, encoding="utf-8-sig")
    print(f"\n  结果已保存 → {out_path}")


if __name__ == "__main__":
    main()
